from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
import random
import time
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl import Workbook
 

# 自動下載ChromeDriver
service = ChromeService(executable_path=ChromeDriverManager().install()) 

# 關閉通知提醒
options = webdriver.ChromeOptions()
prefs = {"profile.default_content_setting_values.notifications": 2}
options.add_experimental_option("prefs", prefs)

# 開啟瀏覽器
chrome = webdriver.Chrome(service=service, options=options)
time.sleep(random.randint(5, 10))

# 進入首頁
chrome.get('https://shopee.tw/buyer/login?next=https%3A%2F%2Fshopee.tw%2F')
time.sleep(random.randint(5, 10))

# 自動輸入帳號和密碼
username = "harry20040723@gmail.com"  
password = "Chiukuanwei8715"  

# HTML元素 name = "loginKey"
username_input = chrome.find_element("name", "loginKey")

# HTML元素 name = "password"
password_input = chrome.find_element("name", "password")

# 模擬使用者輸入帳密
username_input.send_keys(username)
password_input.send_keys(password)

# 模擬按下 Enter 鍵進行登入
password_input.send_keys(Keys.RETURN)
# 按下按鈕送出
#password_input.submit()
time.sleep(5)

# 最大化瀏覽器視窗
chrome.maximize_window()

# 防止身分認證出現
time.sleep(random.randint(15,25))

print('---------- 開始進行爬蟲 ----------')

page = 2 # 爬取2頁
keyword = '運動長褲'
ecode = 'utf-8-sig'

tStart = time.time() # 計時開始

itemid = [] # 商品ID
shopid =[] # 商家ID
itemname = [] # 商品名稱
link = [] # 商品連結
itemprice = [] # 價格
itemsell = [] # 出售數量
itemplace = [] #出貨地

for i in range(int(page)):
    chrome.get('https://shopee.tw/search?keyword=' + keyword + '&page=' + str(i))    
    time.sleep(random.randint(1,5))

    # 滾動頁面
    for scroll in range(0,7):
        chrome.execute_script('window.scrollBy(0,1000)')
        time.sleep(2)

    # 取得商品內容
    for block in chrome.find_elements(by=By.XPATH, value='//*[@data-sqe="item"]'):
        # 將整個網站的Html進行解析
        soup = BeautifulSoup(block.get_attribute('innerHTML'), "html.parser").find('a')
        time.sleep(2)
        # -------------------------------------------- 取得商品連結、商品ID、商家ID --------------------------------------------
        get_link = soup.get('href')
        # rfind() 返回字符串最后一次出现的位置
        # https://shopee.tw/%E3%80%90TPS%E6%9C%89%E4%BB%B6%E8%A4%B2%E5%AD%90%E3%80%91%E7%9E%AC%E6%B6%BC%E5%86%B0%E7%B5%B2%E5%BD%88%E5%8A%9B%E6%9D%9F%E5%8F%A3%E8%A4%B2-
        # %E4%B8%89%E8%89%B2-27-42%E8%85%B0-%E5%9B%9B%E9%9D%A2%E5%BD%88%E5%8A%9B-%E6%8B%89%E9%8D%8A%E5%8F%A3%E8%A2%8B-%E5%90%B8%E6%BF%95%E6%8E%92%E6%B1%97-
        # %E9%80%9F%E4%B9%BE-%E6%B6%BC%E6%84%9F-%E7%94%B7%E5%A5%B3%E5%8F%AF%E7%A9%BF-%E5%8A%A0%E5%A4%A7%E5%B0%BA%E7%A2%BC-i.231200245.22234460888?sp_atk=d896d073-
        # 9d8d-4cd7-9f53-74c46656ac19&xptdk=d896d073-9d8d-4cd7-9f53-74c46656ac19
        # .231200245.22234460888? 商家ID -> 商品ID
        get_itemid = int((get_link[get_link.rfind('.')+1:get_link.rfind('?')]))
        get_shopid = int(get_link[ get_link[:get_link.rfind('.')].rfind('.')+1 :get_link.rfind('.')]) 
        # ---------------------------------------------------------------------------------------------------------------------
        
        # 商品名稱
        get_itemname = ''
        get_parent = soup.find('div',class_="FDn--+")
        get_itemname = get_parent.text
        if len(get_itemname) == 0:
            continue
        
        # 商品當下價格
        get_itemprice = ''
        price_box = soup.find_all('span',class_ = 'ZEgDH9')
        if(len(price_box) > 1):
            get_itemprice = price_box[0].text + '~' + price_box[1].text
        else:
            get_itemprice += price_box[0].text             

        # 已出售數
        get_itemsell = ''
        sell_box = soup.find('div',class_ = 'r6HknA uEPGHT')
        get_itemsell = sell_box.text
        get_itemsell = get_itemsell.replace('已售出','').strip()

        # 出貨地
        get_itemplace = ''
        place_box = soup.find('div',class_ = 'zGGwiV')
        get_itemplace = place_box.text

        # 將資料塞入陣列
        link.append("https://shopee.tw" + get_link)
        itemid.append(get_itemid)
        shopid.append(get_shopid)
        itemname.append(get_itemname)
        itemprice.append(get_itemprice.strip())
        itemsell.append(get_itemsell.strip())
        itemplace.append(get_itemplace)

    time.sleep(random.randint(5,10)) 

# 2023/04/20 先將每頁抓到的商品儲存下來，方便後續追蹤並爬蟲
dic = {
    '商品ID':itemid,
    '賣家ID':shopid,
    '商品名稱':itemname,
    '商品連結':link,
    '價格':itemprice,
    '售出數量':itemsell,
    '出貨地':itemplace     
}

df = pd.DataFrame(dic)

# 新建一個 Excel 工作簿，選擇活動工作表，活動工作表是您在工作簿中正在查看的工作表
workbook = Workbook()
worksheet = workbook.active

# 將 DataFrame 的資料添加到工作表，header=True 則是指包括 DataFrame 的列標題
for row in dataframe_to_rows(df, index=False, header=True):
    worksheet.append(row)

# 將所有欄位的所有單元格靠左對齊，使用 worksheet.iter_rows 方法遍歷工作表中的每一行。min_row=2 指定了從第 2 行開始，而 max_row=worksheet.max_row 則是到工作表的最後一行
for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
    for cell in row:
        cell.alignment = Alignment(horizontal='left')

excel_file_path = r'C:\python-training\實作訓練\檔案存取區\{}_商品資料.xlsx'.format(keyword)
workbook.save(excel_file_path)


# 關閉瀏覽器
chrome.close()
print('successful,Excel檔已存置:',excel_file_path)